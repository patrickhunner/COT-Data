from calendar import c
from csv import writer
from email import header
from json import load
from typing import final
from urllib.request import Request, urlopen
import urllib.request
from wsgiref import headers
from xmlrpc.client import Boolean
import pandas as pd
from datetime import datetime
import openpyxl
from zipfile import ZipFile
import os
from io import BytesIO
import pandas_datareader as pdr
import datetime as dt
import xlsxwriter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

class COT_Update():
    def __init__(self, read):
        self.read = read
        look_at = pd.read_excel(read)
        self.xlsx_df = look_at["File"].to_list()
        self.xlsx_names = look_at["File"].to_list()
        self.tickers = look_at["Ticker"].to_list()
        self.code = look_at["Number"].to_list()
        self.past_years = look_at["Historical"].to_list()
        self.df_headers = look_at["Headers"].to_list()
        self.df_headers = self.df_headers[0:39]
        self.df_analysis_columns = look_at["Analysis"].to_list()
        self.df_analysis_columns = [x for x in self.df_analysis_columns if pd.isnull(x) == False]
        if read == "Commodities_Look_At.xlsx":
            self.file = "Commodities.xlsx"
            self.writer = pd.ExcelWriter(self.file)
            self.num_hist = 8
            self.url = "https://www.cftc.gov/dea/newcot/deafut.txt"
        else:
            self.file = "Financials.xlsx"
            self.writer = pd.ExcelWriter(self.file)
            self.num_hist = 7
            self.url = "https://www.cftc.gov/dea/newcot/FinFutWk.txt"

    def scrape_new(self):
        request_site = Request(self.url, headers={"User-Agent": "Mozilla/5.0"})
        data = urlopen(request_site).read()
        data = pd.read_csv(BytesIO(data), header = None)
        columns = [0, 2, 3, 7, 8, 9, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23]
        data = data.iloc[:,columns]
        for i in range(0,len(data)):
            row = data.iloc[i]
            if str(row[3]) in self.code:
                self.new_data(i,row)

    def new_data(self, num, row):     # adds new row to the top of the spreadsheet
        # cur_df = self.xlsx_df[self.code.index(num)]
        # cur_df.loc[len(cur_df)] = row
        # wb = xlsxwriter.Workbook(self.file)
        print(row)
        return

    def remove_row(self):
        for files in self.csv_names:
            df = pd.read_excel(files, engine = "openpyxl")
            df = df.drop(df.index[0])



analysis = COT_Update("Financials_Look_At.xlsx")
analysis.scrape_new()
