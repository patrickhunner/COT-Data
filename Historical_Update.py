from calendar import c
from csv import writer
from email import header
from json import load
from typing import final
from urllib.request import Request, urlopen
import urllib.request
from wsgiref import headers
from xmlrpc.client import Boolean
import requests
import pandas as pd
from datetime import datetime
import openpyxl
from zipfile import ZipFile
import os
from io import BytesIO
import pandas_datareader as pdr
import datetime as dt
import xlsxwriter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

class COT_Historical():
    def __init__(self, read):
        self.read = read
        look_at = pd.read_excel(read)
        self.xlsx_df = look_at["File"].to_list()
        self.xlsx_names = look_at["File"].to_list()
        self.tickers = look_at["Ticker"].to_list()
        self.code = look_at["Number"].to_list()
        self.past_years = look_at["Historical"].to_list()
        self.df_headers = look_at["Headers"].to_list()
        self.df_headers = self.df_headers[0:39]
        self.df_analysis_columns = look_at["Analysis"].to_list()
        self.df_analysis_columns = [x for x in self.df_analysis_columns if pd.isnull(x) == False]
        if read == "Look_At\Commodities_Look_At.xlsx":
            self.file = "Commodities.xlsx"
            self.writer = pd.ExcelWriter(self.file)
            self.num_hist = 8
        else:
            self.file = "Financials.xlsx"
            self.writer = pd.ExcelWriter(self.file)
            self.num_hist = 7

    def historical_pandas(self):        # convert historical csv to df
        print("Historical Pandas")
        for index in range(0,self.num_hist):
            df = pd.read_csv(self.past_years[index])
            df = df.filter(self.df_headers)
            self.past_years[index] = df
        for index, tick in enumerate(self.tickers):
            if tick != "None" and tick != "Date":
                start = dt.datetime(1980,1,1)
                end = dt.datetime(2022,6,12)
                data = pdr.get_data_yahoo(tick,start,end)
                data = data["Close"]
                self.tickers[index] = data
            else:
                self.tickers[index] = False

    def current_pandas(self, clear):           # convert analysis xlsx to df, clear if requested
        print("Current Pandas")
        if clear == True:               # check if files are empty
            for index in range(0,len(self.xlsx_df)):
                self.xlsx_df[index] = pd.DataFrame(columns = self.df_headers)
        else:
            for index, dfs in enumerate(self.xlsx_df):
                self.xlsx_df[index] = pd.read_excel(self.file, sheet_name = dfs, engine='openpyxl')

    def add_historical(self):           # adds historical data if not already present
        self.historical_pandas()
        self.current_pandas(True)
        print("Add Historical")
        for i in range(0,self.num_hist):
            print(i)
            year = self.past_years[i]
            for index, row in year.iterrows():
                num = (index,row[2])[1]
                if num in self.code:
                    self.add_row(num,row, index)


    def add_row(self, num, row, i):        # adds row to the bottom of the dataframe
        row = row.tolist() + 20*[0]
        date = row[1]
        row[19] = row[4] - row[5]
        row[23] = row[6] - row[7]
        row[27] = row[9] - row[10]
        row[31] = row[11] - row[12]
        add = False
        if type(self.tickers[self.code.index(num)]) != Boolean:
            move = 0
            while True:
                other_date = date.split("/")
                other_date = other_date[2] + "/" + other_date[0] + "/" + other_date[1]
                try:
                    row[35] = self.tickers[self.code.index(num)].loc[other_date]
                    cur_df = self.xlsx_df[self.code.index(num)]
                    cur_df.loc[len(cur_df)] = row
                    break
                except:
                    new_date = date.split("/")
                    if int(new_date[1]) > 27:
                        add = False
                    if int(new_date[1]) < 2:
                        add = True
                    if add == True:
                        new_date[1] = str(int(new_date[1]) + 1)
                        move += 1
                    else:
                        new_date[1] = str(int(new_date[1]) - 1)
                        move -= 1
                    if move <= -5 or move >= 5:
                        self.tickers[self.code.index(num)] = False
                        print(num)
                        break
                    date = new_date[0] + "/" + new_date[1] + "/" + new_date[2]
        return True

    def to_xlsx(self):          # writes all data to update xlsx files
        print("write data to xlsx files")
        for index, files in enumerate(self.xlsx_df):
            ws = self.xlsx_names[index]
            files.to_excel(self.writer, ws, index = False)
        self.writer.save()
            
    def min_max_index(self):        # does what the name says
        # self.current_pandas(False)
        print("min_max_index")
        above = []
        below = []
        for df in self.xlsx_df:     # iterate through each worksheet (commodity/financial)
            if len(df) > 156:
                for i in reversed(range(0,len(df) - 156)):      # iterate bottom to top, starting three years later than earliest
                    row = df.iloc[i]
                    for j in range(0, len(self.df_analysis_columns), 4):        # iterate through the different types of buyers
                        net_column = df.loc[i:i + 156, self.df_analysis_columns[j]]
                        max = net_column.max()
                        min = net_column.min()
                        net = row[self.df_analysis_columns[j]]
                        df.loc[i,self.df_analysis_columns[j + 1]] = max
                        df.loc[i,self.df_analysis_columns[j + 2]] = min
                        df.loc[i,self.df_analysis_columns[j + 3]] = ((net - min) / (max - min)) * 100
        print("above")
        for things in above:
            print(things)
        print("below")
        for things in below:
            print(things)

    # def formatting(self):
    #     wb = load_workbook(filename = self.read)
    #     for sheet in self.xlsx_names:
    #         wb.active = wb[sheet]
    #         print(wb.active)

    #     wb = xlsxwriter.Workbook(self.read)
    #     for sheet in self.xlsx_names:
    #         print(sheet)
    #         ws = wb.get_worksheet_by_name(sheet)
    #         print(ws)

    
    def scrape_new(self):
        # req = Request("https://www.cftc.gov/dea/newcot/f_disagg.txt")
        # data = urlopen(req).read()
        # data = urllib.request.urlopen("https://www.cftc.gov/dea/newcot/f_disagg.txt")
        # data = urllib2.urlopen("https://www.cftc.gov/dea/newcot/f_disagg.txt")
        # data = urllib.request.urlopen("https://www.cftc.gov/dea/newcot/f_disagg.txt")
        url = "https://www.cftc.gov/dea/newcot/deafut.txt"
        request_site = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        data = urlopen(request_site).read()
        data = pd.read_csv(BytesIO(data))
        for index, row in data.iterrows():
            commodity = str((index,row[3])[1])
            if commodity in self.code:
                print(row)
                self.new_data(commodity,row)

    def new_data(self, num, row):     # adds new row to the top of the spreadsheet
        # cur_df = self.xlsx_df[self.code.index(num)]
        # cur_df.loc[len(cur_df)] = row
        wb = xlsxwriter.Workbook(self.file)

    def remove_row(self):
        for files in self.csv_names:
            df = pd.read_excel(files, engine = "openpyxl")
            df = df.drop(df.index[0])

analysis = COT_Historical("Look_At\Financials_Look_At.xlsx")
analysis.add_historical()
analysis.min_max_index()
analysis.to_xlsx()

analysis = COT_Historical("Look_At\Commodities_Look_At.xlsx")
analysis.add_historical()
analysis.min_max_index()
analysis.to_xlsx()